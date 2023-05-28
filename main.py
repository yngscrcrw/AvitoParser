import re
import sys
import base64
import pytesseract
import tracemalloc
import traceback
from io import BytesIO
from PIL import Image
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor


def Process_ad(ad, ws, total_promotions, options, index, page, seleniumwire_options, phone_numbers):
    wait = WebDriverWait(ad, 180)
    title = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.iva-item-titleStep-pdebR h3'))).text
    element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a')))
    product_link = element.get_attribute('href')

    with webdriver.Chrome(options=options) as ad_driver:
        wait = WebDriverWait(ad_driver, 180)
        ad_driver.delete_all_cookies()
        ad_driver.get(product_link)
        views_element = wait.until(EC.presence_of_element_located((By.XPATH, '//span[@data-marker="item-view/total-views"]')))
        views = int(views_element.text.strip().split()[0])
        today_views_element = wait.until(EC.presence_of_element_located((By.XPATH, '//span[@data-marker="item-view/today-views"]')))
        today_views = int(re.search(r'\d+', today_views_element.text).group())

        seller_link = ''
        try:
            seller = ad_driver.find_element(By.CSS_SELECTOR, '.styles-module-size_ms-EVWML').text
            seller_link = ad_driver.find_element(By.CSS_SELECTOR, "a[data-marker='seller-link/link']").get_attribute('href')
        except NoSuchElementException:
            seller = "Нет"

    ad_driver.quit()

    position = index + (page - 1) * 50 + 2
    print(f'{index + (page - 1) * 50} | {title} | {seller} | {total_promotions[index]} | {views} | {today_views} | 'f'{seller_link} | {product_link} | {phone_numbers[index]}')

    ws.cell(row=position, column=1, value=title).hyperlink = product_link
    ws.cell(row=position, column=2, value=seller).hyperlink = seller_link
    ws.cell(row=position, column=3, value=total_promotions[index])
    ws.cell(row=position, column=4, value=views)
    ws.cell(row=position, column=5, value=today_views)
    ws.cell(row=position, column=6, value=phone_numbers[index])

tracemalloc.start()
wb = Workbook()
ws = wb.active
ws.append(['Название', 'Селлер', 'Продвижение', 'Просмотры', 'За сегодня', 'Номер'])
alignment = Alignment(horizontal='center', vertical='center')

options = Options()
proxy_username = 'koWFc3KAzU'
proxy_password = 'QBIbI9L85F'
seleniumwire_options = {
    'proxy': {
        'http': f'http://{proxy_username}:{proxy_password}@85.172.104.207:7157',
        'verify_ssl': False,
    },
}
options.add_argument('--headless')

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.delete_all_cookies()

product = 'репетитор+по+математике'
driver.get("https://www.avito.ru/moskva?cd=1&q=%D0%B2%D0%BE%D0%B4%D0%BE%D0%BE%D1%82%D1%87%D0%B8%D1%81%D1%82%D0%BA%D0%B0")

try:
    pageNum = int(driver.find_element(By.CSS_SELECTOR, '.styles-module-listItem_last-GI_us').text)
except NoSuchElementException:
    print("Проблема с авито")
    sys.exit()

try:
    for page in range(1, 1):
        print('Номер страницы: ' + str(page))
        try:
            ads = driver.find_elements(By.CSS_SELECTOR, ".iva-item-root-_lk9K")
        except NoSuchElementException:
            break
        total_promotions = []
        phone_numbers = []
        total_promotion = ' '
        phone_number = "нет"
        for ad in ads:
            actions = ActionChains(driver)
            wait = WebDriverWait(ad, 50)
            try:
                tooltipElement = ad.find_element(By.CSS_SELECTOR, '.styles-arrow-jfRdd')
                actions.move_to_element(tooltipElement).perform()
                promotions = driver.find_elements(By.CSS_SELECTOR, '.styles-entry-MuP_G')
                for promotion in promotions:
                    promotionElement = promotion.text
                    if promotionElement == "Продвинуто":
                        if re.search("https://www.avito.st/s/common/components/monetization/icons/web/x20",
                                     promotion.find_element(By.CSS_SELECTOR, 'img').get_attribute('src')):
                            promotionElement = "Продвинуто x20"
                        elif re.search("https://www.avito.st/s/common/components/monetization/icons/web/x5",
                                       promotion.find_element(By.CSS_SELECTOR, 'img').get_attribute('src')):
                            promotionElement = "Продвинуто x5"
                        elif re.search("https://www.avito.st/s/common/components/monetization/icons/web/x10",
                                       promotion.find_element(By.CSS_SELECTOR, 'img').get_attribute('src')):
                            promotionElement = "Продвинуто x10"
                        elif re.search("https://www.avito.st/s/common/components/monetization/icons/web/x15",
                                       promotion.find_element(By.CSS_SELECTOR, 'img').get_attribute('src')):
                            promotionElement = "Продвинуто x15"
                        elif re.search("https://www.avito.st/s/common/components/monetization/icons/web/x2.",
                                       promotion.find_element(By.CSS_SELECTOR, 'img').get_attribute('src')):
                            promotionElement = "Продвинуто x2"
                    total_promotion += promotionElement + ' '
            except NoSuchElementException:
                total_promotion = 'нет'
            except TimeoutException:
                total_promotion = 'нет'
            total_promotions.append(total_promotion)
            try:
                button = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'button[data-marker="item-phone-button/undefined"]')))
                actions.move_to_element(button).click().perform()
                phone_img = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.button-phone-image-LkzoU')))
                phone_src = phone_img.get_attribute('src')
                phone_data = base64.b64decode(phone_src.split(',')[1])
                phone_pic = Image.open(BytesIO(phone_data))
                phone_number = pytesseract.image_to_string(phone_pic)
            except NoSuchElementException:
                phone_number = "нет"
            except TimeoutException:
                phone_number = "нет"
            phone_numbers.append(phone_number)
            print(len(total_promotions), ' ', total_promotion, phone_number)
            total_promotion = ' '
            phone_number = "нет"
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            for index, ad in enumerate(ads):
                try:
                    future = executor.submit(Process_ad, ad, ws, total_promotions, options, index, page, seleniumwire_options, phone_numbers)
                    futures.append(future)
                except TimeoutException:
                    continue
            for future in futures:
                result = future.result()

        driver.get(driver.current_url[:-1] + str(page + 1))

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass

    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    for cell in column:
        cell.alignment = alignment

    wb.save(filename='parsing_data.xlsx')
    driver.quit()
except Exception as e:
    traceback.print_exc()
finally:
    wb.save(filename='parsing_data.xlsx')